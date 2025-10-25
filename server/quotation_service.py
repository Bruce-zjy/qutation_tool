#!/usr/bin/env python3
"""
Quotation Service
Handles product search, price calculation, and quotation export (Excel and PDF)
"""

import sys
import json
import pandas as pd
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from datetime import datetime
import os

# Get the directory of this script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCT_DATA_PATH = os.path.join(SCRIPT_DIR, 'product_prices.pkl')

def load_product_data():
    """Load product data from pickle file"""
    return pd.read_pickle(PRODUCT_DATA_PATH)

def search_product(query, limit=10):
    """
    Search for products using fuzzy matching
    Returns a list of matching products with their details
    """
    dataframe = load_product_data()
    temp_df = dataframe.copy()
    temp_df['unique_id'] = temp_df.index

    temp_df['combined_text'] = temp_df['Product'].astype(str).fillna('') + ' ' + \
                                  temp_df['Description'].astype(str).fillna('') + ' ' + \
                                  temp_df['Cut-Off'].astype(str).fillna('') + ' ' + \
                                  temp_df['Pack'].astype(str).fillna('')

    matches = process.extract(query, temp_df['combined_text'], limit=limit*2, scorer=fuzz.token_set_ratio)
    
    all_matches = {}
    for match_str, score, idx in matches:
        if idx not in all_matches or score > all_matches[idx]:
            all_matches[idx] = score

    sorted_matches = sorted(all_matches.items(), key=lambda item: item[1], reverse=True)

    results = []
    for original_idx, score in sorted_matches[:limit]:
        original_row = dataframe.loc[original_idx].copy()
        result = {
            'product': str(original_row['Product']),
            'description': str(original_row['Description']),
            'cutOff': str(original_row['Cut-Off']) if pd.notna(original_row['Cut-Off']) else '',
            'pack': str(original_row['Pack']) if pd.notna(original_row['Pack']) else '',
            'baseUsdFinished': float(original_row['成品_USD']) if pd.notna(original_row['成品_USD']) else 0,
            'baseRmbFinished': float(original_row['成品_RMB']) if pd.notna(original_row['成品_RMB']) else 0,
            'baseUsdBulk': float(original_row['大板_USD']) if pd.notna(original_row['大板_USD']) else 0,
            'baseRmbBulk': float(original_row['大板_RMB']) if pd.notna(original_row['大板_RMB']) else 0,
            'matchScore': int(score)
        }
        results.append(result)
        
    return results

def calculate_quotation_prices(base_usd, add_on_percentage=0.10, exchange_rate=7.1, tax_rate=0.13):
    """
    Calculate final prices based on base USD price
    Returns dict with final_usd and final_rmb
    """
    final_usd = base_usd * (1 + add_on_percentage)
    final_rmb = final_usd * exchange_rate * (1 + tax_rate)
    return {
        'final_usd': round(final_usd, 4),
        'final_rmb': round(final_rmb, 4)
    }

def generate_excel_quotation(quotation_data, output_path):
    """
    Generate Excel quotation file
    quotation_data: dict with keys: quotationNumber, customerName, items (list of dicts)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "QUOTATION"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Quotation info
    ws['A3'] = "Quotation Number:"
    ws['B3'] = quotation_data.get('quotationNumber', '')
    ws['A4'] = "Customer Name:"
    ws['B4'] = quotation_data.get('customerName', '')
    ws['A5'] = "Date:"
    ws['B5'] = datetime.now().strftime('%Y-%m-%d')
    
    # Table headers
    headers = ['No.', 'Product', 'Specimen', 'Format', 'Pack', 'Qty', 
               'Finished USD', 'Finished RMB', 'Bulk USD', 'Bulk RMB']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    items = quotation_data.get('items', [])
    for row_idx, item in enumerate(items, start=8):
        ws.cell(row=row_idx, column=1).value = row_idx - 7
        ws.cell(row=row_idx, column=2).value = item.get('product', '')
        ws.cell(row=row_idx, column=3).value = item.get('specimen', '')
        ws.cell(row=row_idx, column=4).value = item.get('format', '')
        ws.cell(row=row_idx, column=5).value = item.get('pack', '')
        ws.cell(row=row_idx, column=6).value = item.get('quantity', 1)
        ws.cell(row=row_idx, column=7).value = item.get('finalUsdFinished', 0)
        ws.cell(row=row_idx, column=8).value = item.get('finalRmbFinished', 0)
        ws.cell(row=row_idx, column=9).value = item.get('finalUsdBulk', 0) if item.get('finalUsdBulk') else ''
        ws.cell(row=row_idx, column=10).value = item.get('finalRmbBulk', 0) if item.get('finalRmbBulk') else ''
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    wb.save(output_path)
    return output_path

def generate_pdf_quotation(quotation_data, output_path):
    """
    Generate PDF quotation file
    quotation_data: dict with keys: quotationNumber, customerName, items (list of dicts)
    """
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    doc = SimpleDocTemplate(output_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#333333'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("QUOTATION", title_style)
    elements.append(title)
    
    # Quotation info
    info_style = styles['Normal']
    info_data = [
        ['Quotation Number:', quotation_data.get('quotationNumber', '')],
        ['Customer Name:', quotation_data.get('customerName', '')],
        ['Date:', datetime.now().strftime('%Y-%m-%d')]
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    table_data = [['No.', 'Product', 'Specimen', 'Format', 'Pack', 'Qty', 
                   'Finished\nUSD', 'Finished\nRMB', 'Bulk\nUSD', 'Bulk\nRMB']]
    
    items = quotation_data.get('items', [])
    for idx, item in enumerate(items, start=1):
        row = [
            str(idx),
            item.get('product', ''),
            item.get('specimen', ''),
            item.get('format', ''),
            item.get('pack', ''),
            str(item.get('quantity', 1)),
            f"${item.get('finalUsdFinished', 0):.2f}",
            f"¥{item.get('finalRmbFinished', 0):.2f}",
            f"${item.get('finalUsdBulk', 0):.2f}" if item.get('finalUsdBulk') else '',
            f"¥{item.get('finalRmbBulk', 0):.2f}" if item.get('finalRmbBulk') else ''
        ]
        table_data.append(row)
    
    # Create table
    col_widths = [0.4*inch, 1.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.4*inch, 
                  0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    return output_path

def main():
    """Main entry point for CLI usage"""
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No command provided'}))
        sys.exit(1)
    
    command = sys.argv[1]
    
    try:
        if command == 'search':
            if len(sys.argv) < 3:
                print(json.dumps({'error': 'Query parameter required'}))
                sys.exit(1)
            query = sys.argv[2]
            limit = int(sys.argv[3]) if len(sys.argv) > 3 else 10
            results = search_product(query, limit)
            print(json.dumps({'success': True, 'data': results}))
        
        elif command == 'calculate':
            if len(sys.argv) < 3:
                print(json.dumps({'error': 'Base USD price required'}))
                sys.exit(1)
            base_usd = float(sys.argv[2])
            markup = float(sys.argv[3]) if len(sys.argv) > 3 else 0.10
            exchange_rate = float(sys.argv[4]) if len(sys.argv) > 4 else 7.1
            tax_rate = float(sys.argv[5]) if len(sys.argv) > 5 else 0.13
            result = calculate_quotation_prices(base_usd, markup, exchange_rate, tax_rate)
            print(json.dumps({'success': True, 'data': result}))
        
        elif command == 'export_excel':
            if len(sys.argv) < 4:
                print(json.dumps({'error': 'Data and output path required'}))
                sys.exit(1)
            data_json = sys.argv[2]
            output_path = sys.argv[3]
            quotation_data = json.loads(data_json)
            result_path = generate_excel_quotation(quotation_data, output_path)
            print(json.dumps({'success': True, 'path': result_path}))
        
        elif command == 'export_pdf':
            if len(sys.argv) < 4:
                print(json.dumps({'error': 'Data and output path required'}))
                sys.exit(1)
            data_json = sys.argv[2]
            output_path = sys.argv[3]
            quotation_data = json.loads(data_json)
            result_path = generate_pdf_quotation(quotation_data, output_path)
            print(json.dumps({'success': True, 'path': result_path}))
        
        else:
            print(json.dumps({'error': f'Unknown command: {command}'}))
            sys.exit(1)
    
    except Exception as e:
        print(json.dumps({'error': str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()

